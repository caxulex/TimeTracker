"""
Service layer for Payroll Reports generation
"""

from datetime import date, datetime
from decimal import Decimal
from typing import Optional, List
from io import BytesIO
from sqlalchemy import select, and_, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import PayrollPeriod, PayrollEntry, PayrollAdjustment, User, PayRate
from app.schemas.payroll import (
    PayrollReportFilters,
    PayrollSummaryReport,
    UserPayrollReport,
    PayablesDepartmentReport,
    PayrollAdjustmentResponse,
    PeriodStatusEnum
)


class PayrollReportService:
    """Service for generating payroll reports"""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def get_period_summary(self, period_id: int) -> Optional[PayrollSummaryReport]:
        """Generate summary report for a specific period"""
        stmt = select(PayrollPeriod).options(
            selectinload(PayrollPeriod.entries)
        ).where(PayrollPeriod.id == period_id)
        
        result = await self.db.execute(stmt)
        period = result.scalar_one_or_none()
        
        if not period:
            return None
        
        entries = period.entries
        
        return PayrollSummaryReport(
            period_id=period.id,
            period_name=period.name,
            start_date=period.start_date,
            end_date=period.end_date,
            status=period.status,
            total_employees=len(entries),
            total_regular_hours=sum(e.regular_hours for e in entries),
            total_overtime_hours=sum(e.overtime_hours for e in entries),
            total_gross_amount=sum(e.gross_amount for e in entries),
            total_adjustments=sum(e.adjustments_amount for e in entries),
            total_net_amount=sum(e.net_amount for e in entries)
        )
    
    async def get_user_payroll_report(
        self,
        user_id: int,
        period_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> List[UserPayrollReport]:
        """Generate payroll report for a specific user"""
        conditions = [PayrollEntry.user_id == user_id]
        
        if period_id:
            conditions.append(PayrollEntry.payroll_period_id == period_id)
        
        stmt = select(PayrollEntry).options(
            selectinload(PayrollEntry.user),
            selectinload(PayrollEntry.period),
            selectinload(PayrollEntry.adjustments).selectinload(PayrollAdjustment.creator)
        ).where(and_(*conditions))
        
        if start_date or end_date:
            stmt = stmt.join(PayrollPeriod)
            if start_date:
                stmt = stmt.where(PayrollPeriod.start_date >= start_date)
            if end_date:
                stmt = stmt.where(PayrollPeriod.end_date <= end_date)
        
        stmt = stmt.order_by(PayrollEntry.created_at.desc())
        
        result = await self.db.execute(stmt)
        entries = result.scalars().all()
        
        reports = []
        for entry in entries:
            adjustments = [
                PayrollAdjustmentResponse(
                    id=adj.id,
                    payroll_entry_id=adj.payroll_entry_id,
                    adjustment_type=adj.adjustment_type,
                    description=adj.description,
                    amount=adj.amount,
                    created_by=adj.created_by,
                    created_at=adj.created_at,
                    created_by_name=adj.creator.name if adj.creator else None
                )
                for adj in entry.adjustments
            ]
            
            reports.append(UserPayrollReport(
                user_id=entry.user_id,
                user_name=entry.user.name,
                user_email=entry.user.email,
                period_name=entry.period.name,
                start_date=entry.period.start_date,
                end_date=entry.period.end_date,
                regular_hours=entry.regular_hours,
                overtime_hours=entry.overtime_hours,
                regular_rate=entry.regular_rate,
                overtime_rate=entry.overtime_rate,
                gross_amount=entry.gross_amount,
                adjustments=adjustments,
                adjustments_total=entry.adjustments_amount,
                net_amount=entry.net_amount
            ))
        
        return reports
    
    async def get_payables_report(
        self,
        filters: PayrollReportFilters
    ) -> PayablesDepartmentReport:
        """Generate comprehensive report for payables department"""
        conditions = []
        
        if filters.period_id:
            conditions.append(PayrollPeriod.id == filters.period_id)
        if filters.status:
            conditions.append(PayrollPeriod.status == filters.status.value)
        if filters.period_type:
            conditions.append(PayrollPeriod.period_type == filters.period_type.value)
        if filters.start_date:
            conditions.append(PayrollPeriod.start_date >= filters.start_date)
        if filters.end_date:
            conditions.append(PayrollPeriod.end_date <= filters.end_date)
        
        stmt = select(PayrollPeriod).options(
            selectinload(PayrollPeriod.entries).selectinload(PayrollEntry.user),
            selectinload(PayrollPeriod.entries).selectinload(PayrollEntry.adjustments).selectinload(PayrollAdjustment.creator)
        )
        
        if conditions:
            stmt = stmt.where(and_(*conditions))
        
        stmt = stmt.order_by(PayrollPeriod.start_date.desc())
        
        result = await self.db.execute(stmt)
        periods = result.scalars().all()
        
        # Aggregate all entries
        all_entries = []
        total_regular_hours = Decimal("0.00")
        total_overtime_hours = Decimal("0.00")
        total_gross = Decimal("0.00")
        total_adjustments = Decimal("0.00")
        total_net = Decimal("0.00")
        
        for period in periods:
            for entry in period.entries:
                if filters.user_id and entry.user_id != filters.user_id:
                    continue
                # Filter by company_id if specified
                if filters.company_id is not None and entry.user and entry.user.company_id != filters.company_id:
                    continue
                
                # Get user's rate type from their active pay rate
                rate_type = None
                pay_rate_stmt = select(PayRate).where(
                    and_(
                        PayRate.user_id == entry.user_id,
                        PayRate.is_active == True
                    )
                ).order_by(PayRate.effective_from.desc()).limit(1)
                pay_rate_result = await self.db.execute(pay_rate_stmt)
                pay_rate = pay_rate_result.scalar_one_or_none()
                if pay_rate:
                    rate_type = pay_rate.rate_type
                
                adjustments = [
                    PayrollAdjustmentResponse(
                        id=adj.id,
                        payroll_entry_id=adj.payroll_entry_id,
                        adjustment_type=adj.adjustment_type,
                        description=adj.description,
                        amount=adj.amount,
                        created_by=adj.created_by,
                        created_at=adj.created_at,
                        created_by_name=adj.creator.name if adj.creator else None
                    )
                    for adj in entry.adjustments
                ]
                
                all_entries.append(UserPayrollReport(
                    user_id=entry.user_id,
                    user_name=entry.user.name,
                    user_email=entry.user.email,
                    rate_type=rate_type,
                    period_name=period.name,
                    start_date=period.start_date,
                    end_date=period.end_date,
                    regular_hours=entry.regular_hours,
                    overtime_hours=entry.overtime_hours,
                    regular_rate=entry.regular_rate,
                    overtime_rate=entry.overtime_rate,
                    gross_amount=entry.gross_amount,
                    adjustments=adjustments,
                    adjustments_total=entry.adjustments_amount,
                    net_amount=entry.net_amount
                ))
                
                total_regular_hours += entry.regular_hours
                total_overtime_hours += entry.overtime_hours
                total_gross += entry.gross_amount
                total_adjustments += entry.adjustments_amount
                total_net += entry.net_amount
        
        # Create summary
        summary = PayrollSummaryReport(
            period_id=filters.period_id or 0,
            period_name=f"Multiple Periods ({len(periods)})" if len(periods) > 1 else (periods[0].name if periods else "No Data"),
            start_date=filters.start_date or (min(p.start_date for p in periods) if periods else date.today()),
            end_date=filters.end_date or (max(p.end_date for p in periods) if periods else date.today()),
            status=filters.status.value if filters.status else "mixed",
            total_employees=len(set(e.user_id for e in all_entries)),
            total_regular_hours=total_regular_hours,
            total_overtime_hours=total_overtime_hours,
            total_gross_amount=total_gross,
            total_adjustments=total_adjustments,
            total_net_amount=total_net
        )
        
        # Determine report period string
        if filters.start_date and filters.end_date:
            report_period = f"{filters.start_date} to {filters.end_date}"
        elif filters.period_id and periods:
            report_period = f"{periods[0].start_date} to {periods[0].end_date}"
        else:
            report_period = "All Time"
        
        return PayablesDepartmentReport(
            report_generated_at=datetime.utcnow(),
            report_period=report_period,
            filters_applied=filters,
            summary=summary,
            entries=all_entries
        )
    
    async def export_to_csv(self, report: PayablesDepartmentReport) -> str:
        """Export payables report to CSV format"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Report Generated", report.report_generated_at.isoformat(),
            "Period", report.report_period
        ])
        writer.writerow([])
        
        # Summary
        writer.writerow(["SUMMARY"])
        writer.writerow(["Total Employees", report.summary.total_employees])
        writer.writerow(["Total Regular Hours", str(report.summary.total_regular_hours)])
        writer.writerow(["Total Overtime Hours", str(report.summary.total_overtime_hours)])
        writer.writerow(["Total Gross Amount", str(report.summary.total_gross_amount)])
        writer.writerow(["Total Adjustments", str(report.summary.total_adjustments)])
        writer.writerow(["Total Net Amount", str(report.summary.total_net_amount)])
        writer.writerow([])
        
        # Detail header
        writer.writerow([
            "User ID", "User Name", "Email", "Period",
            "Regular Hours", "Overtime Hours",
            "Regular Rate", "Overtime Rate",
            "Gross Amount", "Adjustments", "Net Amount"
        ])
        
        # Detail rows
        for entry in report.entries:
            writer.writerow([
                entry.user_id,
                entry.user_name,
                entry.user_email,
                entry.period_name,
                str(entry.regular_hours),
                str(entry.overtime_hours),
                str(entry.regular_rate),
                str(entry.overtime_rate),
                str(entry.gross_amount),
                str(entry.adjustments_total),
                str(entry.net_amount)
            ])
        
        return output.getvalue()
    
    async def export_to_excel(self, report: PayablesDepartmentReport) -> bytes:
        """Export payables report to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        if ws_summary is None:
            ws_summary = wb.create_sheet('Summary')
        else:
            ws_summary.title = 'Summary'
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws_summary["A1"] = "Payroll Report"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A2"] = f"Generated: {report.report_generated_at.isoformat()}"
        ws_summary["A3"] = f"Period: {report.report_period}"
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Employees", report.summary.total_employees],
            ["Total Regular Hours", float(report.summary.total_regular_hours)],
            ["Total Overtime Hours", float(report.summary.total_overtime_hours)],
            ["Total Gross Amount", float(report.summary.total_gross_amount)],
            ["Total Adjustments", float(report.summary.total_adjustments)],
            ["Total Net Amount", float(report.summary.total_net_amount)],
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Detail Sheet
        ws_detail = wb.create_sheet("Detail")
        
        headers = [
            "User ID", "User Name", "Email", "Period",
            "Regular Hours", "Overtime Hours",
            "Regular Rate", "Overtime Rate",
            "Gross Amount", "Adjustments", "Net Amount"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, entry in enumerate(report.entries, start=2):
            ws_detail.cell(row=row_idx, column=1, value=entry.user_id)
            ws_detail.cell(row=row_idx, column=2, value=entry.user_name)
            ws_detail.cell(row=row_idx, column=3, value=entry.user_email)
            ws_detail.cell(row=row_idx, column=4, value=entry.period_name)
            ws_detail.cell(row=row_idx, column=5, value=float(entry.regular_hours))
            ws_detail.cell(row=row_idx, column=6, value=float(entry.overtime_hours))
            ws_detail.cell(row=row_idx, column=7, value=float(entry.regular_rate))
            ws_detail.cell(row=row_idx, column=8, value=float(entry.overtime_rate))
            ws_detail.cell(row=row_idx, column=9, value=float(entry.gross_amount))
            ws_detail.cell(row=row_idx, column=10, value=float(entry.adjustments_total))
            ws_detail.cell(row=row_idx, column=11, value=float(entry.net_amount))
        
        # Adjust column widths
        for ws in [ws_summary, ws_detail]:
            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
