"""
Export router for generating PDF and Excel reports
"""

from datetime import datetime, date, timedelta
from typing import Optional, List
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_

from app.database import get_db
from app.models import User, Project, Task, TimeEntry, Team, TeamMember
from app.dependencies import get_current_active_user

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

router = APIRouter()


def format_duration(seconds: int) -> str:
    """Format duration in seconds to HH:MM:SS"""
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


async def get_user_time_entries(
    db: AsyncSession,
    user: User,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    project_id: Optional[int] = None,
    user_id: Optional[int] = None
) -> List[dict]:
    """Get time entries with filtering"""
    query = select(TimeEntry, Project.name.label("project_name"), Task.name.label("task_name")).outerjoin(
        Project, TimeEntry.project_id == Project.id
    ).outerjoin(
        Task, TimeEntry.task_id == Task.id
    )

    # Apply user filter based on role
    if user.role not in ["super_admin", "admin"]:
        if user_id and user_id != user.id:
            raise HTTPException(status_code=403, detail="Access denied")
        query = query.where(TimeEntry.user_id == user.id)
    elif user_id:
        query = query.where(TimeEntry.user_id == user_id)

    # Date filters
    if start_date:
        query = query.where(TimeEntry.start_time >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(TimeEntry.start_time <= datetime.combine(end_date, datetime.max.time()))
    if project_id:
        query = query.where(TimeEntry.project_id == project_id)

    query = query.order_by(TimeEntry.start_time.desc())
    result = await db.execute(query)
    rows = result.all()

    entries = []
    for row in rows:
        entry = row[0]
        duration = 0
        if entry.end_time and entry.start_time:
            duration = int((entry.end_time - entry.start_time).total_seconds())
        
        entries.append({
            "id": entry.id,
            "date": entry.start_time.strftime("%Y-%m-%d"),
            "start_time": entry.start_time.strftime("%H:%M:%S"),
            "end_time": entry.end_time.strftime("%H:%M:%S") if entry.end_time else "Running",
            "duration": format_duration(duration),
            "duration_seconds": duration,
            "project": row[1] or "No Project",
            "task": row[2] or "No Task",
            "description": entry.description or ""
        })

    return entries


@router.get("/excel")
async def export_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    project_id: Optional[int] = None,
    user_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export time entries to Excel format"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available. Install openpyxl.")

    entries = await get_user_time_entries(db, current_user, start_date, end_date, project_id, user_id)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Entries"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Headers
    headers = ["Date", "Start Time", "End Time", "Duration", "Project", "Task", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry["date"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=entry["start_time"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=entry["end_time"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=entry["duration"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=entry["project"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=entry["task"]).border = thin_border
        ws.cell(row=row_idx, column=7, value=entry["description"]).border = thin_border

    # Summary row
    total_seconds = sum(e["duration_seconds"] for e in entries)
    summary_row = len(entries) + 3
    ws.cell(row=summary_row, column=3, value="Total:").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=format_duration(total_seconds)).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 30

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"time_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/pdf")
async def export_pdf(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    project_id: Optional[int] = None,
    user_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export time entries to PDF format"""
    if not PDF_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF export not available. Install reportlab.")

    entries = await get_user_time_entries(db, current_user, start_date, end_date, project_id, user_id)

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=20,
        textColor=colors.HexColor("#4F46E5")
    )
    elements.append(Paragraph("Time Tracking Report", title_style))
    
    # Date range subtitle
    date_range = ""
    if start_date and end_date:
        date_range = f"Period: {start_date} to {end_date}"
    elif start_date:
        date_range = f"From: {start_date}"
    elif end_date:
        date_range = f"Until: {end_date}"
    else:
        date_range = "All Time"
    
    elements.append(Paragraph(date_range, styles["Normal"]))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Table data
    table_data = [["Date", "Start", "End", "Duration", "Project", "Task"]]
    for entry in entries:
        table_data.append([
            entry["date"],
            entry["start_time"],
            entry["end_time"],
            entry["duration"],
            entry["project"][:15] + "..." if len(entry["project"]) > 15 else entry["project"],
            entry["task"][:15] + "..." if len(entry["task"]) > 15 else entry["task"]
        ])

    # Summary
    total_seconds = sum(e["duration_seconds"] for e in entries)
    table_data.append(["", "", "Total:", format_duration(total_seconds), "", ""])

    # Create table
    table = Table(table_data, colWidths=[70, 55, 55, 60, 100, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -2), colors.white),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -2), 0.5, colors.grey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    filename = f"time_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/csv")
async def export_csv(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    project_id: Optional[int] = None,
    user_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export time entries to CSV format"""
    import csv
    
    entries = await get_user_time_entries(db, current_user, start_date, end_date, project_id, user_id)

    buffer = BytesIO()
    
    # Write CSV to buffer
    import io
    text_buffer = io.StringIO()
    writer = csv.writer(text_buffer)
    
    # Header
    writer.writerow(["Date", "Start Time", "End Time", "Duration", "Project", "Task", "Description"])
    
    # Data
    for entry in entries:
        writer.writerow([
            entry["date"],
            entry["start_time"],
            entry["end_time"],
            entry["duration"],
            entry["project"],
            entry["task"],
            entry["description"]
        ])
    
    # Convert to bytes
    buffer.write(text_buffer.getvalue().encode("utf-8"))
    buffer.seek(0)

    filename = f"time_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
