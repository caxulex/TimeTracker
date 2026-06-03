"""
Reports and analytics router
"""

import logging
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import (
    FILTER_NULL_COMPANY,
    apply_company_filter,
    get_company_filter,
    get_company_timezone,
    get_current_active_user,
)
from app.models import Project, ProjectTeam, Task, Team, TeamMember, TimeEntry, User
from app.services.duration_service import calculate_entry_duration_for_period
from app.services.email_log_utils import log_email_failed, log_email_sent
from app.utils.timer_elapsed import compute_display_elapsed_seconds
from app.utils.timewindow import (
    day_bounds,
    local_today,
    month_bounds,
    now_utc,
    range_bounds,
    week_bounds,
)

router = APIRouter()
logger = logging.getLogger(__name__)


class ProjectSummary(BaseModel):
    project_id: int
    project_name: str
    total_seconds: int
    total_hours: float
    entry_count: int
    billable_amount: Optional[float] = None
    budget_hours: Optional[float] = None
    budget_used_percent: Optional[float] = None


class UserSummary(BaseModel):
    user_id: int
    user_name: str
    total_seconds: int
    total_hours: float
    entry_count: int


class TaskSummary(BaseModel):
    task_id: int
    task_name: str
    project_name: str
    total_seconds: int
    total_hours: float
    status: str


class DailySummary(BaseModel):
    date: date
    total_seconds: int
    total_hours: float
    entry_count: int


class WeeklySummary(BaseModel):
    week_start: date
    week_end: date
    total_seconds: int
    total_hours: float
    daily_breakdown: List[DailySummary]


class DashboardStats(BaseModel):
    today_seconds: int
    today_hours: float
    week_seconds: int
    week_hours: float
    month_seconds: int
    month_hours: float
    active_projects: int
    pending_tasks: int
    running_timer: bool


class TimeReport(BaseModel):
    start_date: date
    end_date: date
    total_seconds: int
    total_hours: float
    total_entries: int
    by_project: List[ProjectSummary]
    by_user: List[UserSummary]
    by_day: List[DailySummary]


# Team Timesheet Report Models
class TeamTimesheetUserEntry(BaseModel):
    """Entry for a single user's day in the timesheet"""
    date: date
    seconds: int
    formatted: str  # HH:MM format


class TeamTimesheetUser(BaseModel):
    """Single user row in the Team Timesheet"""
    user_id: int
    user_name: str
    role: str
    daily_hours: List[TeamTimesheetUserEntry]
    total_seconds: int
    total_formatted: str  # HH:MM format


class TeamTimesheetDayTotal(BaseModel):
    """Total for a single day (column total)"""
    date: date
    seconds: int
    formatted: str  # HH:MM format


class TeamTimesheetReport(BaseModel):
    """Complete Team Timesheet report"""
    start_date: date
    end_date: date
    dates: List[date]  # List of all dates in range for column headers
    users: List[TeamTimesheetUser]  # User rows with their daily hours
    daily_totals: List[TeamTimesheetDayTotal]  # Column totals for each day
    grand_total_seconds: int
    grand_total_formatted: str  # HH:MM format


def format_seconds_to_hhmm(seconds: int) -> str:
    """Format seconds to HH:MM format (e.g., 3661 -> '1:01')"""
    if seconds <= 0:
        return "0:00"
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return f"{hours}:{minutes:02d}"


def calculate_entry_duration(entry: TimeEntry, now: datetime) -> int:
    """Calculate duration for a time entry, including running timers.

    For running timers, returns the pause-aware live elapsed (frozen while
    on break, otherwise wall-clock minus accumulated pause_seconds). For
    closed entries, returns the stored ``duration_seconds`` which is itself
    pause-corrected at /stop and /switch time (see PR #31).
    """
    if entry.end_time is None:
        # Active timer - use the shared pause-aware helper so reports agree
        # with the live "Who's Working Now" / timer widget displays.
        return compute_display_elapsed_seconds(entry, now=now)
    return entry.duration_seconds or 0


# ``calculate_entry_duration_for_period`` is imported above from
# ``app.services.duration_service`` (consolidated from a duplicate that
# previously lived both here and in ``app/ai/services/reporting_service.py``).


@router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """Get personal dashboard statistics for the authenticated user.

    Returns ONLY the current user's time entries regardless of role. Admins
    who want company-wide views should use /api/reports/admin/dashboard.

    B6/B7: All day/week/month bounds are computed in the company's local
    timezone, then converted to UTC for SQL filters. This is what the user
    perceives as "today" in their tenancy.
    """
    now = now_utc()
    today_local = local_today(tz)
    today_start, today_end = day_bounds(today_local, tz)
    week_start, week_end = week_bounds(today_local, tz)
    month_start, month_end = month_bounds(today_local, tz)

    # Multi-tenancy: get company filter (still used for active_projects scope)
    company_id = get_company_filter(current_user)

    # Personal scope: always filter to the current user's own entries.
    user_filter = TimeEntry.user_id == current_user.id

    # Today's time - fetch entries that OVERLAP with today (started before today end AND ended after today start or still running)
    # This includes: entries that started today, entries from yesterday still running, entries spanning midnight
    today_query = select(TimeEntry).where(
        TimeEntry.start_time < today_end,  # Started before today ends
        (TimeEntry.end_time >= today_start) | (TimeEntry.end_time.is_(None)),  # Ended after today started OR still running
        user_filter,
    )
    today_result = await db.execute(today_query)
    today_entries = today_result.scalars().all()
    # Calculate only the portion that falls within today
    today_seconds = sum(calculate_entry_duration_for_period(e, today_start, today_end, now) for e in today_entries)

    # This week's time - fetch entries that overlap with this week
    week_query = select(TimeEntry).where(
        TimeEntry.start_time < week_end,
        (TimeEntry.end_time >= week_start) | (TimeEntry.end_time.is_(None)),
        user_filter,
    )
    week_result = await db.execute(week_query)
    week_entries = week_result.scalars().all()
    week_seconds = sum(calculate_entry_duration_for_period(e, week_start, week_end, now) for e in week_entries)

    # This month's time - fetch entries that overlap with this month
    month_query = select(TimeEntry).where(
        TimeEntry.start_time < month_end,
        (TimeEntry.end_time >= month_start) | (TimeEntry.end_time.is_(None)),
        user_filter,
    )
    month_result = await db.execute(month_query)
    month_entries = month_result.scalars().all()
    month_seconds = sum(calculate_entry_duration_for_period(e, month_start, month_end, now) for e in month_entries)

    # Active projects (user has access to, within company)
    project_query = select(func.count(Project.id)).join(Team, Project.team_id == Team.id).where(
        Project.is_archived == False,
        Team.deleted_at.is_(None),
    )
    if current_user.role not in ["super_admin", "admin", "company_admin"]:
        user_teams = select(TeamMember.team_id).where(TeamMember.user_id == current_user.id)
        # Visibility rule for regular users in reports: count projects
        # visible via primary ownership OR project_teams association.
        project_query = project_query.where(
            or_(
                Project.team_id.in_(user_teams),
                Project.id.in_(
                    select(ProjectTeam.project_id).where(
                        ProjectTeam.team_id.in_(user_teams)
                    )
                ),
            )
        )
    if company_id is None:
        pass  # Super admin sees all
    elif company_id == FILTER_NULL_COMPANY:
        project_query = project_query.where(Team.company_id.is_(None))
    else:
        project_query = project_query.where(Team.company_id == company_id)
    active_projects_result = await db.execute(project_query)
    active_projects = active_projects_result.scalar() or 0

    # Pending tasks assigned to user
    pending_tasks = 0  # Simplified for now

    # Check for running timer
    running_timer = any(e.end_time is None for e in today_entries)

    return DashboardStats(
        today_seconds=today_seconds,
        today_hours=round(today_seconds / 3600, 2),
        week_seconds=week_seconds,
        week_hours=round(week_seconds / 3600, 2),
        month_seconds=month_seconds,
        month_hours=round(month_seconds / 3600, 2),
        active_projects=active_projects,
        pending_tasks=pending_tasks,
        running_timer=running_timer
    )


@router.get("/weekly", response_model=WeeklySummary)
async def get_weekly_summary(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    week_offset: int = Query(0, ge=-52, le=0, description="Weeks ago (0 = current week)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """Get personal time summary for the authenticated user over an arbitrary date range.

    Returns ONLY the current user's time entries regardless of role. Admins
    who want company-wide views should use /api/reports/admin/dashboard.

    Date range behavior:
    - If both ``start_date`` and ``end_date`` are provided, the window spans that
      caller-controlled range (inclusive). The ``daily_breakdown`` will contain
      one entry per day in the range.
    - If only ``start_date`` is provided, the window is the 7 days starting at
      ``start_date`` (backwards compatible with the original weekly behavior).
    - If neither is provided, the window is the current week in the tenant's
      timezone, optionally offset by ``week_offset``.

    The range is capped at 366 days to prevent excessively large responses.
    """
    now = now_utc()
    today_local = local_today(tz)

    # Use start_date if provided, otherwise calculate from week_offset (in local tz)
    if start_date:
        week_start = start_date
    else:
        week_start = today_local - timedelta(days=today_local.weekday()) - timedelta(weeks=abs(week_offset))

    # Honor caller-supplied end_date when present; otherwise keep the historical
    # 7-day window so existing callers continue to work unchanged.
    if end_date is not None:
        week_end = end_date
    else:
        week_end = week_start + timedelta(days=6)

    if week_end < week_start:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="end_date must be on or after start_date",
        )

    range_days = (week_end - week_start).days + 1
    if range_days > 366:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Date range too large (max 366 days)",
        )

    # Half-open bounds in tenant local time (B7)
    start_datetime, end_datetime = range_bounds(week_start, week_end, tz)

    # Personal scope: always filter to the current user's own entries.
    user_filter = TimeEntry.user_id == current_user.id

    # Fetch entries that OVERLAP with this window (not just started within)
    entries_query = select(TimeEntry).where(
        TimeEntry.start_time < end_datetime,  # Started before window ends
        (TimeEntry.end_time >= start_datetime) | (TimeEntry.end_time.is_(None)),  # Ended after window started OR still running
        user_filter,
    )
    entries_result = await db.execute(entries_query)
    all_entries = entries_result.scalars().all()

    # Calculate total seconds for the window using period overlap
    total_seconds = sum(calculate_entry_duration_for_period(e, start_datetime, end_datetime, now) for e in all_entries)

    # Daily breakdown - calculate overlap for each day in the range
    daily_breakdown = []
    for i in range(range_days):
        day = week_start + timedelta(days=i)
        day_start, day_end = day_bounds(day, tz)

        # Calculate seconds for this day from ALL entries that overlap with this day
        day_seconds = sum(calculate_entry_duration_for_period(e, day_start, day_end, now) for e in all_entries)
        # Count entries that have any overlap with this day
        day_count = sum(1 for e in all_entries if calculate_entry_duration_for_period(e, day_start, day_end, now) > 0)

        daily_breakdown.append(DailySummary(
            date=day,
            total_seconds=day_seconds,
            total_hours=round(day_seconds / 3600, 2),
            entry_count=day_count
        ))

    return WeeklySummary(
        week_start=week_start,
        week_end=week_end,
        total_seconds=total_seconds,
        total_hours=round(total_seconds / 3600, 2),
        daily_breakdown=daily_breakdown
    )


@router.get("/by-project", response_model=List[ProjectSummary])
async def get_time_by_project(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """Get personal time summary grouped by project for the authenticated user.

    Returns ONLY the current user's time entries regardless of role. Admins
    who want company-wide views should use /api/reports/admin/dashboard.
    """
    now = now_utc()
    today_local = local_today(tz)

    # Default to current local month
    if not start_date:
        start_date = today_local.replace(day=1)
    if not end_date:
        end_date = today_local

    start_datetime, end_datetime = range_bounds(start_date, end_date, tz)

    # Personal scope: always filter to the current user's own entries.
    user_filter = TimeEntry.user_id == current_user.id

    # Fetch entries that OVERLAP with the period (not just started within)
    query_filters = [
        TimeEntry.start_time < end_datetime,  # Started before period ends
        (TimeEntry.end_time >= start_datetime) | (TimeEntry.end_time.is_(None)),  # Ended after period started OR still running
        user_filter,
    ]

    result = await db.execute(
        select(TimeEntry, Project.name)
        .outerjoin(Project, TimeEntry.project_id == Project.id)
        .where(*query_filters)
    )

    # Group by project and calculate totals - only count time within the period
    project_data: Dict[int, Dict[str, Any]] = defaultdict(lambda: {"name": "", "seconds": 0, "count": 0})

    for entry, project_name in result.all():
        pid = entry.project_id or 0  # Group meeting entries (NULL project) under key 0
        project_data[pid]["name"] = project_name or "Meeting"
        # Calculate only the portion that falls within the requested period
        entry_seconds = calculate_entry_duration_for_period(entry, start_datetime, end_datetime, now)
        if entry_seconds > 0:
            project_data[pid]["seconds"] += entry_seconds
            project_data[pid]["count"] += 1

    summaries = []
    for project_id, data in sorted(project_data.items(), key=lambda x: x[1]["seconds"], reverse=True):
        total_seconds = data["seconds"]
        total_hours = round(total_seconds / 3600, 2)

        summaries.append(ProjectSummary(
            project_id=project_id,
            project_name=data["name"],
            total_seconds=total_seconds,
            total_hours=total_hours,
            entry_count=data["count"],
            billable_amount=None,
            budget_hours=None,
            budget_used_percent=None
        ))

    return summaries

@router.get("/by-task", response_model=List[TaskSummary])
async def get_time_by_task(
    project_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """Get time summary grouped by task"""
    now = now_utc()
    today_local = local_today(tz)

    # Default to current local month
    if not start_date:
        start_date = today_local.replace(day=1)
    if not end_date:
        end_date = today_local

    start_datetime, end_datetime = range_bounds(start_date, end_date, tz)

    # Fetch entries that OVERLAP with the period
    query = (
        select(TimeEntry, Task.name.label("task_name"), Task.status, Project.name.label("project_name"))
        .join(Task, TimeEntry.task_id == Task.id)
        .outerjoin(Project, TimeEntry.project_id == Project.id)
        .where(
            TimeEntry.user_id == current_user.id,
            TimeEntry.task_id != None,
            TimeEntry.start_time < end_datetime,
            (TimeEntry.end_time >= start_datetime) | (TimeEntry.end_time.is_(None))
        )
    )

    if project_id:
        query = query.where(TimeEntry.project_id == project_id)

    result = await db.execute(query)

    # Group by task and calculate totals - only count time within period
    task_data: Dict[int, Dict[str, Any]] = defaultdict(lambda: {"task_name": "", "project_name": "", "status": "", "seconds": 0})

    for entry, task_name, task_status, project_name in result.all():
        tid = entry.task_id
        task_data[tid]["task_name"] = task_name
        task_data[tid]["project_name"] = project_name
        task_data[tid]["status"] = task_status
        entry_seconds = calculate_entry_duration_for_period(entry, start_datetime, end_datetime, now)
        if entry_seconds > 0:
            task_data[tid]["seconds"] += entry_seconds

    summaries = []
    for task_id, data in sorted(task_data.items(), key=lambda x: x[1]["seconds"], reverse=True):
        total_seconds = data["seconds"]
        summaries.append(TaskSummary(
            task_id=task_id,
            task_name=data["task_name"],
            project_name=data["project_name"],
            total_seconds=total_seconds,
            total_hours=round(total_seconds / 3600, 2),
            status=data["status"]
        ))

    return summaries


@router.get("/team", response_model=TimeReport)
async def get_team_report(
    team_id: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """Get team time report (team admin/owner only, filtered by company for multi-tenancy)"""
    # Multi-tenancy: verify team belongs to user's company
    company_id = get_company_filter(current_user)
    team_query = select(Team).where(Team.id == team_id)
    team_query = apply_company_filter(team_query, Team.company_id, company_id)
    team_query = team_query.where(Team.deleted_at.is_(None))
    team_result = await db.execute(team_query)
    team = team_result.scalar_one_or_none()
    if not team:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Team not found")

    # Check team access
    if current_user.role not in ["super_admin", "admin", "company_admin"]:
        member_check = await db.execute(
            select(TeamMember).where(
                TeamMember.team_id == team_id,
                TeamMember.user_id == current_user.id,
                TeamMember.role.in_(["owner", "admin"])
            )
        )
        if not member_check.scalar_one_or_none():
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Team admin access required")

    # Default to current local month
    if not start_date:
        start_date = local_today(tz).replace(day=1)
    if not end_date:
        end_date = local_today(tz)

    now = now_utc()
    start_datetime, end_datetime = range_bounds(start_date, end_date, tz)

    # Get team members
    team_members = select(TeamMember.user_id).where(TeamMember.team_id == team_id)

    # Team report visibility includes projects where the team is primary
    # owner and projects explicitly shared via project_teams.
    team_projects = select(Project.id).where(
        or_(
            Project.team_id == team_id,
            Project.id.in_(
                select(ProjectTeam.project_id).where(ProjectTeam.team_id == team_id)
            ),
        )
    )

    # Fetch all entries that OVERLAP with the period (instead of using SQL aggregates)
    entries_query = (
        select(TimeEntry, Project.name.label("project_name"), User.name.label("user_name"))
        .outerjoin(Project, TimeEntry.project_id == Project.id)
        .join(User, TimeEntry.user_id == User.id)
        .where(
            TimeEntry.user_id.in_(team_members),
            TimeEntry.project_id.in_(team_projects),
            TimeEntry.start_time < end_datetime,
            (TimeEntry.end_time >= start_datetime) | (TimeEntry.end_time.is_(None))
        )
    )
    entries_result = await db.execute(entries_query)
    all_entries = entries_result.all()

    # Calculate totals with proper period overlap
    total_seconds = 0
    total_entries = 0
    project_data: Dict[int, Dict[str, Any]] = defaultdict(lambda: {"name": "", "seconds": 0, "count": 0})
    user_data: Dict[int, Dict[str, Any]] = defaultdict(lambda: {"name": "", "seconds": 0, "count": 0})

    for entry, project_name, user_name in all_entries:
        entry_seconds = calculate_entry_duration_for_period(entry, start_datetime, end_datetime, now)
        if entry_seconds > 0:
            total_seconds += entry_seconds
            total_entries += 1
            project_data[entry.project_id]["name"] = project_name
            project_data[entry.project_id]["seconds"] += entry_seconds
            project_data[entry.project_id]["count"] += 1
            user_data[entry.user_id]["name"] = user_name
            user_data[entry.user_id]["seconds"] += entry_seconds
            user_data[entry.user_id]["count"] += 1

    # Build by_project list
    by_project = []
    for pid, data in sorted(project_data.items(), key=lambda x: x[1]["seconds"], reverse=True):
        seconds = data["seconds"]
        by_project.append(ProjectSummary(
            project_id=pid,
            project_name=data["name"],
            total_seconds=seconds,
            total_hours=round(seconds / 3600, 2),
            entry_count=data["count"],
            billable_amount=None,
            budget_hours=None,
            budget_used_percent=None
        ))

    # Build by_user list
    by_user = []
    for uid, data in sorted(user_data.items(), key=lambda x: x[1]["seconds"], reverse=True):
        seconds = data["seconds"]
        by_user.append(UserSummary(
            user_id=uid,
            user_name=data["name"],
            total_seconds=seconds,
            total_hours=round(seconds / 3600, 2),
            entry_count=data["count"]
        ))

    # By day - calculate overlap for each day
    by_day = []
    current_date = start_date
    while current_date <= end_date:
        day_start, day_end = day_bounds(current_date, tz)

        day_seconds = 0
        day_count = 0
        for entry, _, _ in all_entries:
            entry_day_seconds = calculate_entry_duration_for_period(entry, day_start, day_end, now)
            if entry_day_seconds > 0:
                day_seconds += entry_day_seconds
                day_count += 1

        by_day.append(DailySummary(
            date=current_date,
            total_seconds=day_seconds,
            total_hours=round(day_seconds / 3600, 2),
            entry_count=day_count
        ))

        current_date += timedelta(days=1)

    return TimeReport(
        start_date=start_date,
        end_date=end_date,
        total_seconds=total_seconds,
        total_hours=round(total_seconds / 3600, 2),
        total_entries=total_entries,
        by_project=by_project,
        by_user=by_user,
        by_day=by_day
    )


@router.get("/export")
async def export_time_entries(
    start_date: date,
    end_date: date,
    project_id: Optional[int] = None,
    format: str = Query("json", pattern="^(json|csv)$"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """Export time entries (JSON or CSV format)"""
    start_datetime, end_datetime = range_bounds(start_date, end_date, tz)

    # Fetch entries that OVERLAP with the export period
    query = (
        select(
            TimeEntry,
            Project.name.label("project_name"),
            Task.name.label("task_name"),
            User.name.label("user_name")
        )
        .outerjoin(Project, TimeEntry.project_id == Project.id)
        .outerjoin(Task, TimeEntry.task_id == Task.id)
        .join(User, TimeEntry.user_id == User.id)
        .where(
            TimeEntry.user_id == current_user.id,
            TimeEntry.start_time < end_datetime,
            (TimeEntry.end_time >= start_datetime) | (TimeEntry.end_time.is_(None))
        )
        .order_by(TimeEntry.start_time.desc())
    )

    if project_id:
        query = query.where(TimeEntry.project_id == project_id)

    result = await db.execute(query)
    rows = result.all()

    if format == "csv":
        import csv
        import io

        from fastapi.responses import StreamingResponse

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Date", "Start Time", "End Time", "Duration (hours)",
            "Project", "Task", "Description"
        ])

        for row in rows:
            entry = row[0]
            writer.writerow([
                entry.start_time.date().isoformat(),
                entry.start_time.time().isoformat(),
                entry.end_time.time().isoformat() if entry.end_time else "",
                round(entry.duration_seconds / 3600, 2) if entry.duration_seconds else "",
                row.project_name,
                row.task_name or "",
                entry.description or ""
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=time_entries_{start_date}_{end_date}.csv"}
        )

    # JSON format
    entries = []
    for row in rows:
        entry = row[0]
        entries.append({
            "id": entry.id,
            "date": entry.start_time.date().isoformat(),
            "start_time": entry.start_time.isoformat(),
            "end_time": entry.end_time.isoformat() if entry.end_time else None,
            "duration_hours": round(entry.duration_seconds / 3600, 2) if entry.duration_seconds else None,
            "project": row.project_name,
            "task": row.task_name,
            "description": entry.description
        })

    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "total_entries": len(entries),
        "entries": entries
    }




class AdminDashboardStats(BaseModel):
    total_today_seconds: int
    total_today_hours: float
    total_week_seconds: int
    total_week_hours: float
    total_month_seconds: int
    total_month_hours: float
    active_users_today: int
    active_projects: int
    running_timers: int
    by_user: List[UserSummary]


class TeamAnalytics(BaseModel):
    team_id: int
    team_name: str
    member_count: int
    total_today_seconds: int
    total_today_hours: float
    total_week_seconds: int
    total_week_hours: float
    total_month_seconds: int
    total_month_hours: float
    active_members_today: int
    running_timers: int
    top_performers: List[UserSummary]  # Top 3 this week


class IndividualUserMetrics(BaseModel):
    user_id: int
    user_name: str
    user_email: str
    role: str
    teams: List[str]
    # Time metrics
    today_seconds: int
    today_hours: float
    week_seconds: int
    week_hours: float
    month_seconds: int
    month_hours: float
    # Activity metrics
    total_entries: int
    active_days_this_month: int
    avg_hours_per_day: float
    current_timer_running: bool
    # Project breakdown
    projects: List[ProjectSummary]
    # Recent activity
    last_activity: Optional[datetime] = None


class UserAnalyticsRangeMetrics(BaseModel):
    """Aggregated metrics for a single user over a custom date range.

    All totals are computed server-side via SQL aggregation. This endpoint
    exists so views like StaffPage / StaffDetailPage do not have to fetch a
    paginated list of time entries and reduce client-side (which silently
    truncates at the /api/time page-size cap).
    """
    user_id: int
    user_name: str
    start_date: date
    end_date: date
    total_seconds: int
    total_hours: float
    total_entries: int
    days_worked: int
    project_count: int
    avg_hours_per_entry: float
    projects: List[ProjectSummary]


@router.get("/admin/dashboard", response_model=AdminDashboardStats)
async def get_admin_dashboard(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    '''Get admin dashboard with all team members time (admin and super_admin, filtered by company)'''
    if current_user.role not in ["super_admin", "admin", "company_admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    now = now_utc()
    today_local = local_today(tz)
    today_start, today_end = day_bounds(today_local, tz)
    week_start, week_end = week_bounds(today_local, tz)
    month_start, month_end = month_bounds(today_local, tz)

    # Multi-tenancy: get company filter
    company_id = get_company_filter(current_user)

    # Build company user filter
    if company_id is None:
        user_filter = True  # Platform super_admin sees all
    elif company_id == FILTER_NULL_COMPANY:
        company_users_subq = select(User.id).where(User.company_id.is_(None))
        user_filter = TimeEntry.user_id.in_(company_users_subq)
    else:
        company_users_subq = select(User.id).where(User.company_id == company_id)
        user_filter = TimeEntry.user_id.in_(company_users_subq)

    # Total time today - entries that OVERLAP with today
    today_query = select(TimeEntry).where(
        TimeEntry.start_time < today_end,
        (TimeEntry.end_time >= today_start) | (TimeEntry.end_time.is_(None))
    )
    if user_filter is not True:
        today_query = today_query.where(user_filter)
    today_entries_result = await db.execute(today_query)
    today_entries = today_entries_result.scalars().all()

    logger.info(f"Found {len(today_entries)} time entries overlapping with today")

    # Calculate only the portion that falls within today
    total_today = sum(calculate_entry_duration_for_period(e, today_start, today_end, now) for e in today_entries)

    logger.info(f"FINAL total_today={total_today}")

    # Total time this week - entries that overlap with this week
    week_query = select(TimeEntry).where(
        TimeEntry.start_time < week_end,
        (TimeEntry.end_time >= week_start) | (TimeEntry.end_time.is_(None))
    )
    if user_filter is not True:
        week_query = week_query.where(user_filter)
    week_entries_result = await db.execute(week_query)
    week_entries = week_entries_result.scalars().all()

    total_week = sum(calculate_entry_duration_for_period(e, week_start, week_end, now) for e in week_entries)

    # Total time this month - entries that overlap with this month
    month_query = select(TimeEntry).where(
        TimeEntry.start_time < month_end,
        (TimeEntry.end_time >= month_start) | (TimeEntry.end_time.is_(None))
    )
    if user_filter is not True:
        month_query = month_query.where(user_filter)
    month_entries_result = await db.execute(month_query)
    month_entries = month_entries_result.scalars().all()

    total_month = sum(calculate_entry_duration_for_period(e, month_start, month_end, now) for e in month_entries)

    # Active users today - count distinct users from entries overlapping with today
    active_users = len(set(e.user_id for e in today_entries if calculate_entry_duration_for_period(e, today_start, today_end, now) > 0))

    # Active projects (within company)
    project_query = select(func.count(Project.id)).join(Team, Project.team_id == Team.id).where(Project.is_archived == False)
    project_query = project_query.where(Team.deleted_at.is_(None))
    if company_id is None:
        pass  # Super admin sees all
    elif company_id == FILTER_NULL_COMPANY:
        project_query = project_query.where(Team.company_id.is_(None))
    else:
        project_query = project_query.where(Team.company_id == company_id)
    active_projects_result = await db.execute(project_query)
    active_projects = active_projects_result.scalar() or 0

    # Running timers count (within company)
    running_query = select(func.count(TimeEntry.id)).where(TimeEntry.end_time.is_(None))
    if user_filter is not True:
        running_query = running_query.where(user_filter)
    running_result = await db.execute(running_query)
    running_timers = running_result.scalar() or 0

    # Time by user today - use the already fetched today_entries with period calculation
    user_totals = {}
    for entry in today_entries:
        user_id = entry.user_id

        if user_id not in user_totals:
            # Need to fetch user name
            user_totals[user_id] = {
                "user_name": None,  # Will be populated below
                "total_seconds": 0,
                "entry_count": 0
            }

        # Calculate only the portion that falls within today
        entry_seconds = calculate_entry_duration_for_period(entry, today_start, today_end, now)
        if entry_seconds > 0:
            user_totals[user_id]["total_seconds"] += entry_seconds
            user_totals[user_id]["entry_count"] += 1

    # Fetch user names for users with entries
    if user_totals:
        user_names_query = select(User.id, User.name).where(User.id.in_(list(user_totals.keys())))
        user_names_result = await db.execute(user_names_query)
        for row in user_names_result.all():
            if row.id in user_totals:
                user_totals[row.id]["user_name"] = row.name

    by_user = []
    for user_id, data in sorted(user_totals.items(), key=lambda x: x[1]["total_seconds"], reverse=True):
        by_user.append(UserSummary(
            user_id=user_id,
            user_name=data["user_name"],
            total_seconds=data["total_seconds"],
            total_hours=round(data["total_seconds"] / 3600, 2),
            entry_count=data["entry_count"]
        ))

    return AdminDashboardStats(
        total_today_seconds=total_today,
        total_today_hours=round(total_today / 3600, 2),
        total_week_seconds=total_week,
        total_week_hours=round(total_week / 3600, 2),
        total_month_seconds=total_month,
        total_month_hours=round(total_month / 3600, 2),
        active_users_today=active_users,
        active_projects=active_projects,
        running_timers=running_timers,
        by_user=by_user
    )


@router.get("/admin/teams", response_model=List[TeamAnalytics])
async def get_team_analytics(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    '''Get analytics for all teams (admin and super_admin only, filtered by company)'''
    if current_user.role not in ["super_admin", "admin", "company_admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    now = now_utc()
    today_local = local_today(tz)
    today_start, _ = day_bounds(today_local, tz)
    week_start, _ = week_bounds(today_local, tz)
    month_start, _ = month_bounds(today_local, tz)

    # Multi-tenancy: filter teams by company
    company_id = get_company_filter(current_user)
    teams_query = select(Team)
    teams_query = apply_company_filter(teams_query, Team.company_id, company_id)
    teams_query = teams_query.where(Team.deleted_at.is_(None))
    teams_result = await db.execute(teams_query)
    teams = teams_result.scalars().all()

    team_analytics = []

    for team in teams:
        # Get team members
        members_result = await db.execute(
            select(TeamMember.user_id, User.name)
            .join(User, TeamMember.user_id == User.id)
            .where(TeamMember.team_id == team.id)
        )
        members = members_result.all()
        member_ids = [m.user_id for m in members]

        if not member_ids:
            # Empty team, skip
            continue

        # Total time today (team members) - including active timers
        today_entries_result = await db.execute(
            select(TimeEntry)
            .where(
                and_(
                    TimeEntry.start_time >= today_start,
                    TimeEntry.user_id.in_(member_ids)
                )
            )
        )
        today_entries = today_entries_result.scalars().all()

        total_today = 0
        for entry in today_entries:
            if entry.end_time is None:
                start = entry.start_time
                if start.tzinfo is None:
                    start = start.replace(tzinfo=timezone.utc)
                total_today += int((now - start).total_seconds())
            else:
                total_today += (entry.duration_seconds or 0)

        # Total time this week (team members)
        week_entries_result = await db.execute(
            select(TimeEntry)
            .where(
                and_(
                    TimeEntry.start_time >= week_start,
                    TimeEntry.user_id.in_(member_ids)
                )
            )
        )
        week_entries = week_entries_result.scalars().all()

        total_week = 0
        for entry in week_entries:
            if entry.end_time is None:
                start = entry.start_time
                if start.tzinfo is None:
                    start = start.replace(tzinfo=timezone.utc)
                total_week += int((now - start).total_seconds())
            else:
                total_week += (entry.duration_seconds or 0)

        # Total time this month (team members)
        month_entries_result = await db.execute(
            select(TimeEntry)
            .where(
                and_(
                    TimeEntry.start_time >= month_start,
                    TimeEntry.user_id.in_(member_ids)
                )
            )
        )
        month_entries = month_entries_result.scalars().all()

        total_month = 0
        for entry in month_entries:
            if entry.end_time is None:
                start = entry.start_time
                if start.tzinfo is None:
                    start = start.replace(tzinfo=timezone.utc)
                total_month += int((now - start).total_seconds())
            else:
                total_month += (entry.duration_seconds or 0)

        # Active members today
        active_members_result = await db.execute(
            select(func.count(func.distinct(TimeEntry.user_id)))
            .where(
                and_(
                    TimeEntry.start_time >= today_start,
                    TimeEntry.user_id.in_(member_ids)
                )
            )
        )
        active_members = active_members_result.scalar() or 0

        # Running timers count
        running_result = await db.execute(
            select(func.count(TimeEntry.id))
            .where(
                and_(
                    TimeEntry.end_time == None,
                    TimeEntry.user_id.in_(member_ids)
                )
            )
        )
        running_timers = running_result.scalar() or 0

        # Top performers this week (top 3)
        user_result = await db.execute(
            select(
                TimeEntry.user_id,
                User.name,
                TimeEntry.duration_seconds,
                TimeEntry.start_time,
                TimeEntry.end_time
            )
            .join(User, TimeEntry.user_id == User.id)
            .where(
                and_(
                    TimeEntry.start_time >= week_start,
                    TimeEntry.user_id.in_(member_ids)
                )
            )
        )

        # Aggregate by user
        user_totals = {}
        for row in user_result.all():
            user_id = row.user_id
            user_name = row.name

            if user_id not in user_totals:
                user_totals[user_id] = {
                    "user_name": user_name,
                    "total_seconds": 0,
                    "entry_count": 0
                }

            if row.end_time is None:
                start = row.start_time
                if start.tzinfo is None:
                    start = start.replace(tzinfo=timezone.utc)
                user_totals[user_id]["total_seconds"] += int((now - start).total_seconds())
            else:
                user_totals[user_id]["total_seconds"] += (row.duration_seconds or 0)

            user_totals[user_id]["entry_count"] += 1

        # Get top 3 performers
        top_performers = []
        for user_id, data in sorted(user_totals.items(), key=lambda x: x[1]["total_seconds"], reverse=True)[:3]:
            top_performers.append(UserSummary(
                user_id=user_id,
                user_name=data["user_name"],
                total_seconds=data["total_seconds"],
                total_hours=round(data["total_seconds"] / 3600, 2),
                entry_count=data["entry_count"]
            ))

        team_analytics.append(TeamAnalytics(
            team_id=team.id,
            team_name=team.name,
            member_count=len(member_ids),
            total_today_seconds=total_today,
            total_today_hours=round(total_today / 3600, 2),
            total_week_seconds=total_week,
            total_week_hours=round(total_week / 3600, 2),
            total_month_seconds=total_month,
            total_month_hours=round(total_month / 3600, 2),
            active_members_today=active_members,
            running_timers=running_timers,
            top_performers=top_performers
        ))

    return team_analytics


@router.get(
    "/admin/users/{user_id}/analytics",
    response_model=UserAnalyticsRangeMetrics,
)
async def get_user_analytics_range(
    user_id: int,
    start_date: date = Query(..., description="Start of period (inclusive, local date)"),
    end_date: date = Query(..., description="End of period (inclusive, local date)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """Server-side aggregated analytics for a single user over a custom date range.

    Returns totals + project breakdown computed by SQL (no client-side reduce
    of a paginated time-entries list, which was capped at 100 rows by /api/time
    and silently undercounted active users — see PR-A #36 and the May 22
    reports-audit).

    Used by StaffPage analytics, StaffPage TimeTrackingModal, StaffDetailPage.
    Admin/company_admin/super_admin only; filtered by company for multi-tenancy.
    Pause-aware via calculate_entry_duration_for_period (PR #34 helpers).
    """
    if current_user.role not in ["super_admin", "admin", "company_admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required",
        )

    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="end_date must be on or after start_date",
        )

    # Multi-tenancy: filter user by company
    company_id = get_company_filter(current_user)
    user_query = select(User).where(User.id == user_id)
    user_query = apply_company_filter(user_query, User.company_id, company_id)

    user_result = await db.execute(user_query)
    user = user_result.scalar_one_or_none()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found",
        )

    # Resolve the local-day window in UTC for filtering TimeEntry.start_time.
    period_start, period_end = range_bounds(start_date, end_date, tz)
    now = now_utc()

    # Pull every entry that could overlap the window. We include entries whose
    # start_time is before period_end (running timers can still contribute) and
    # whose end_time is None or >= period_start. We still rely on
    # calculate_entry_duration_for_period to clip overlap and apply pause math.
    entries_query = (
        select(TimeEntry, Project.name)
        .outerjoin(Project, TimeEntry.project_id == Project.id)
        .where(
            and_(
                TimeEntry.user_id == user_id,
                TimeEntry.start_time < period_end,
            )
        )
    )
    entries_result = await db.execute(entries_query)
    rows = entries_result.all()

    total_seconds = 0
    total_entries = 0
    days_worked: set = set()
    project_totals: Dict[int, Dict[str, Any]] = {}

    for entry, project_name in rows:
        # Skip closed entries that ended before the window starts.
        entry_end = entry.end_time
        if entry_end is not None:
            if entry_end.tzinfo is None:
                entry_end = entry_end.replace(tzinfo=timezone.utc)
            if entry_end < period_start:
                continue

        seconds = calculate_entry_duration_for_period(
            entry, period_start, period_end, now
        )
        if seconds <= 0:
            continue

        total_seconds += seconds
        total_entries += 1

        # Local day bucket for days_worked
        entry_start = entry.start_time
        if entry_start.tzinfo is None:
            entry_start = entry_start.replace(tzinfo=timezone.utc)
        # Clip to window start so an entry that began before the window only
        # counts the days it actually overlapped.
        bucket_day = max(entry_start, period_start).date()
        days_worked.add(bucket_day)

        pid = entry.project_id or 0
        pname = project_name or "Meeting"
        bucket = project_totals.get(pid)
        if bucket is None:
            bucket = {"name": pname, "seconds": 0, "entries": 0}
            project_totals[pid] = bucket
        bucket["seconds"] += seconds
        bucket["entries"] += 1

    projects = [
        ProjectSummary(
            project_id=pid,
            project_name=data["name"],
            total_seconds=data["seconds"],
            total_hours=round(data["seconds"] / 3600, 2),
            entry_count=data["entries"],
        )
        for pid, data in sorted(
            project_totals.items(), key=lambda x: x[1]["seconds"], reverse=True
        )
    ]

    avg_hours_per_entry = (
        round((total_seconds / 3600) / total_entries, 2) if total_entries else 0.0
    )

    return UserAnalyticsRangeMetrics(
        user_id=user.id,
        user_name=user.name,
        start_date=start_date,
        end_date=end_date,
        total_seconds=total_seconds,
        total_hours=round(total_seconds / 3600, 2),
        total_entries=total_entries,
        days_worked=len(days_worked),
        project_count=sum(1 for pid in project_totals if pid != 0),
        avg_hours_per_entry=avg_hours_per_entry,
        projects=projects,
    )


@router.get("/admin/users/{user_id}", response_model=IndividualUserMetrics)
async def get_user_metrics(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    '''Get detailed metrics for a specific user (admin and super_admin only, filtered by company)'''
    if current_user.role not in ["super_admin", "admin", "company_admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    # Multi-tenancy: filter user by company
    company_id = get_company_filter(current_user)
    user_query = select(User).where(User.id == user_id)
    user_query = apply_company_filter(user_query, User.company_id, company_id)

    user_result = await db.execute(user_query)
    user = user_result.scalar_one_or_none()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )

    now = now_utc()
    today_local = local_today(tz)
    today_start, _ = day_bounds(today_local, tz)
    week_start, _ = week_bounds(today_local, tz)
    month_start, _ = month_bounds(today_local, tz)

    # Get user's teams
    teams_result = await db.execute(
        select(Team.name)
        .join(TeamMember, Team.id == TeamMember.team_id)
        .where(TeamMember.user_id == user_id)
    )
    teams = [t[0] for t in teams_result.all()]

    # Time today
    today_entries_result = await db.execute(
        select(TimeEntry)
        .where(
            and_(
                TimeEntry.start_time >= today_start,
                TimeEntry.user_id == user_id
            )
        )
    )
    today_entries = today_entries_result.scalars().all()

    today_seconds = 0
    for entry in today_entries:
        if entry.end_time is None:
            start = entry.start_time
            if start.tzinfo is None:
                start = start.replace(tzinfo=timezone.utc)
            today_seconds += int((now - start).total_seconds())
        else:
            today_seconds += (entry.duration_seconds or 0)

    # Time this week
    week_entries_result = await db.execute(
        select(TimeEntry)
        .where(
            and_(
                TimeEntry.start_time >= week_start,
                TimeEntry.user_id == user_id
            )
        )
    )
    week_entries = week_entries_result.scalars().all()

    week_seconds = 0
    for entry in week_entries:
        if entry.end_time is None:
            start = entry.start_time
            if start.tzinfo is None:
                start = start.replace(tzinfo=timezone.utc)
            week_seconds += int((now - start).total_seconds())
        else:
            week_seconds += (entry.duration_seconds or 0)

    # Time this month
    month_entries_result = await db.execute(
        select(TimeEntry)
        .where(
            and_(
                TimeEntry.start_time >= month_start,
                TimeEntry.user_id == user_id
            )
        )
    )
    month_entries = month_entries_result.scalars().all()

    month_seconds = 0
    for entry in month_entries:
        if entry.end_time is None:
            start = entry.start_time
            if start.tzinfo is None:
                start = start.replace(tzinfo=timezone.utc)
            month_seconds += int((now - start).total_seconds())
        else:
            month_seconds += (entry.duration_seconds or 0)

    # Total entries count
    total_entries_result = await db.execute(
        select(func.count(TimeEntry.id))
        .where(TimeEntry.user_id == user_id)
    )
    total_entries = total_entries_result.scalar() or 0

    # Active days this month
    active_days_result = await db.execute(
        select(func.count(func.distinct(func.date(TimeEntry.start_time))))
        .where(
            and_(
                TimeEntry.start_time >= month_start,
                TimeEntry.user_id == user_id
            )
        )
    )
    active_days = active_days_result.scalar() or 0

    # Average hours per day (this month)
    avg_hours_per_day = round(month_seconds / 3600 / max(active_days, 1), 2)

    # Check for running timer
    running_result = await db.execute(
        select(TimeEntry)
        .where(
            and_(
                TimeEntry.user_id == user_id,
                TimeEntry.end_time == None
            )
        )
    )
    current_timer_running = running_result.scalar_one_or_none() is not None

    # Project breakdown (this month)
    project_result = await db.execute(
        select(
            TimeEntry.project_id,
            Project.name,
            TimeEntry.duration_seconds,
            TimeEntry.start_time,
            TimeEntry.end_time
        )
        .outerjoin(Project, TimeEntry.project_id == Project.id)
        .where(
            and_(
                TimeEntry.start_time >= month_start,
                TimeEntry.user_id == user_id
            )
        )
    )

    project_totals = {}
    for row in project_result.all():
        project_id = row.project_id or 0  # Group meeting entries under key 0
        project_name = row.name or "Meeting"

        if project_id not in project_totals:
            project_totals[project_id] = {
                "project_name": project_name,
                "total_seconds": 0,
                "entry_count": 0
            }

        if row.end_time is None:
            start = row.start_time
            if start.tzinfo is None:
                start = start.replace(tzinfo=timezone.utc)
            project_totals[project_id]["total_seconds"] += int((now - start).total_seconds())
        else:
            project_totals[project_id]["total_seconds"] += (row.duration_seconds or 0)

        project_totals[project_id]["entry_count"] += 1

    projects = []
    for project_id, data in sorted(project_totals.items(), key=lambda x: x[1]["total_seconds"], reverse=True):
        projects.append(ProjectSummary(
            project_id=project_id,
            project_name=data["project_name"],
            total_seconds=data["total_seconds"],
            total_hours=round(data["total_seconds"] / 3600, 2),
            entry_count=data["entry_count"]
        ))

    # Last activity
    last_activity_result = await db.execute(
        select(TimeEntry.start_time)
        .where(TimeEntry.user_id == user_id)
        .order_by(TimeEntry.start_time.desc())
        .limit(1)
    )
    last_activity_row = last_activity_result.scalar_one_or_none()

    return IndividualUserMetrics(
        user_id=user.id,
        user_name=user.name,
        user_email=user.email,
        role=user.role,
        teams=teams,
        today_seconds=today_seconds,
        today_hours=round(today_seconds / 3600, 2),
        week_seconds=week_seconds,
        week_hours=round(week_seconds / 3600, 2),
        month_seconds=month_seconds,
        month_hours=round(month_seconds / 3600, 2),
        total_entries=total_entries,
        active_days_this_month=active_days,
        avg_hours_per_day=avg_hours_per_day,
        current_timer_running=current_timer_running,
        projects=projects,
        last_activity=last_activity_row
    )


@router.get("/admin/users")
async def get_all_users_summary(
    period: str = Query("week", regex="^(today|week|month)$"),
    page: Optional[int] = Query(None, ge=1, description="Page number (1-indexed). Omit for all results."),
    page_size: Optional[int] = Query(None, ge=1, le=200, description="Results per page. Omit for all results."),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    '''Get summary of all users sorted by time tracked (admin and super_admin only, filtered by company).
    Supports optional pagination via page/page_size query params.
    If omitted, returns full list for backward compatibility.'''
    if current_user.role not in ["super_admin", "admin", "company_admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    now = now_utc()
    today_local = local_today(tz)
    today_start, _ = day_bounds(today_local, tz)
    week_start, _ = week_bounds(today_local, tz)
    month_start, _ = month_bounds(today_local, tz)

    # Determine start time based on period
    if period == "today":
        start_time = today_start
    elif period == "week":
        start_time = week_start
    else:  # month
        start_time = month_start

    # Multi-tenancy: filter by company
    company_id = get_company_filter(current_user)

    # First, get ALL active users (not just those with time entries)
    all_users_query = select(User.id, User.name).where(User.is_active == True)
    all_users_query = apply_company_filter(all_users_query, User.company_id, company_id)
    all_users_result = await db.execute(all_users_query)

    # Initialize all users with zero time
    user_totals = {}
    for row in all_users_result.all():
        user_totals[row.id] = {
            "user_name": row.name or f"User {row.id}",
            "total_seconds": 0,
            "entry_count": 0
        }

    # Get time entries for the period (filtered by company)
    entries_query = select(
        TimeEntry.user_id,
        User.name,
        TimeEntry.duration_seconds,
        TimeEntry.start_time,
        TimeEntry.end_time
    ).join(User, TimeEntry.user_id == User.id).where(TimeEntry.start_time >= start_time)

    entries_query = apply_company_filter(entries_query, User.company_id, company_id)

    entries_result = await db.execute(entries_query)

    for row in entries_result.all():
        user_id = row.user_id

        # Only update if user exists in our list (active users)
        if user_id not in user_totals:
            continue

        if row.end_time is None:
            start = row.start_time
            if start.tzinfo is None:
                start = start.replace(tzinfo=timezone.utc)
            user_totals[user_id]["total_seconds"] += int((now - start).total_seconds())
        else:
            user_totals[user_id]["total_seconds"] += (row.duration_seconds or 0)

        user_totals[user_id]["entry_count"] += 1

    # Sort by total time descending
    users_summary = []
    for user_id, data in sorted(user_totals.items(), key=lambda x: x[1]["total_seconds"], reverse=True):
        users_summary.append(UserSummary(
            user_id=user_id,
            user_name=data["user_name"],
            total_seconds=data["total_seconds"],
            total_hours=round(data["total_seconds"] / 3600, 2),
            entry_count=data["entry_count"]
        ))

    # If no pagination params, return full list (backward compatible)
    if page is None or page_size is None:
        return users_summary

    # Paginated response
    total = len(users_summary)
    total_pages = max(1, (total + page_size - 1) // page_size)

    if page > total_pages and total > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Page {page} exceeds total pages ({total_pages})"
        )

    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    page_data = users_summary[start_idx:end_idx]

    return {
        "data": page_data,
        "total": total,
        "page": page,
        "page_size": page_size,
        "has_next": page < total_pages,
        "has_prev": page > 1,
        "total_pages": total_pages,
    }


@router.get("/team-timesheet", response_model=TeamTimesheetReport)
async def get_team_timesheet(
    start_date: date,
    end_date: date,
    team_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """
    Get Team Timesheet report - shows hours worked per user per day in a grid format.

    This report displays:
    - Rows: Each team member with their name and role
    - Columns: Each day in the date range
    - Cells: Hours worked in HH:MM format (0:00 or dash if none)
    - Horizontal totals: User's total across all days
    - Vertical totals: Team total for each day
    - Grand total: Total hours for entire team across date range

    Args:
        start_date: Start of the reporting period (inclusive)
        end_date: End of the reporting period (inclusive)
        team_id: Optional filter by specific team (admins see all teams otherwise)

    Returns:
        TeamTimesheetReport with user rows, daily columns, and totals
    """
    # Only admins/managers can view team timesheet
    if current_user.role not in ["super_admin", "admin", "company_admin", "manager"]:
        # Check if user is a team admin for the specified team
        if team_id:
            member_check = await db.execute(
                select(TeamMember).where(
                    TeamMember.team_id == team_id,
                    TeamMember.user_id == current_user.id,
                    TeamMember.role.in_(["owner", "admin"])
                )
            )
            if not member_check.scalar_one_or_none():
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Admin or team admin access required"
                )
        else:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required to view all teams"
            )

    now = now_utc()
    start_datetime, end_datetime = range_bounds(start_date, end_date, tz)

    # Multi-tenancy: filter by company
    company_id = get_company_filter(current_user)

    # Build user query based on team filter and company
    if team_id:
        # Specific team selected
        team_query = select(Team).where(Team.id == team_id)
        team_query = apply_company_filter(team_query, Team.company_id, company_id)
        team_query = team_query.where(Team.deleted_at.is_(None))
        team_result = await db.execute(team_query)
        team = team_result.scalar_one_or_none()
        if not team:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Team not found")

        # Get team members
        users_query = (
            select(User.id, User.name, User.role)
            .join(TeamMember, TeamMember.user_id == User.id)
            .where(TeamMember.team_id == team_id, User.is_active == True)
            .distinct()
        )
    else:
        # All users in company (admin view)
        users_query = select(User.id, User.name, User.role).where(User.is_active == True)
        users_query = apply_company_filter(users_query, User.company_id, company_id)

    users_result = await db.execute(users_query)
    users = users_result.all()

    if not users:
        # Return empty report
        dates_list = []
        current = start_date
        while current <= end_date:
            dates_list.append(current)
            current += timedelta(days=1)

        return TeamTimesheetReport(
            start_date=start_date,
            end_date=end_date,
            dates=dates_list,
            users=[],
            daily_totals=[TeamTimesheetDayTotal(date=d, seconds=0, formatted="0:00") for d in dates_list],
            grand_total_seconds=0,
            grand_total_formatted="0:00"
        )

    user_ids = [u.id for u in users]
    user_info = {u.id: {"name": u.name, "role": u.role or "employee"} for u in users}

    # Fetch all time entries for these users in the date range
    entries_query = (
        select(TimeEntry)
        .where(
            TimeEntry.user_id.in_(user_ids),
            TimeEntry.start_time < end_datetime,
            (TimeEntry.end_time >= start_datetime) | (TimeEntry.end_time.is_(None))
        )
    )
    entries_result = await db.execute(entries_query)
    all_entries = entries_result.scalars().all()

    # Generate list of dates in range
    dates_list = []
    current = start_date
    while current <= end_date:
        dates_list.append(current)
        current += timedelta(days=1)

    # Build user-day matrix: user_id -> date -> seconds
    user_day_seconds = defaultdict(lambda: defaultdict(int))

    for entry in all_entries:
        user_id = entry.user_id
        # Calculate seconds for each day this entry overlaps
        for day in dates_list:
            day_start, day_end = day_bounds(day, tz)
            day_seconds = calculate_entry_duration_for_period(entry, day_start, day_end, now)
            if day_seconds > 0:
                user_day_seconds[user_id][day] += day_seconds

    # Build response data
    timesheet_users = []
    daily_totals_seconds = defaultdict(int)
    grand_total = 0

    # Sort users by name for consistent ordering
    sorted_users = sorted(users, key=lambda u: u.name.lower())

    for user in sorted_users:
        user_id = user.id
        info = user_info[user_id]
        user_total = 0
        daily_hours = []

        for day in dates_list:
            seconds = user_day_seconds[user_id].get(day, 0)
            user_total += seconds
            daily_totals_seconds[day] += seconds

            daily_hours.append(TeamTimesheetUserEntry(
                date=day,
                seconds=seconds,
                formatted=format_seconds_to_hhmm(seconds) if seconds > 0 else "-"
            ))

        grand_total += user_total

        timesheet_users.append(TeamTimesheetUser(
            user_id=user_id,
            user_name=info["name"],
            role=info["role"].replace("_", " ").title(),  # Format role nicely
            daily_hours=daily_hours,
            total_seconds=user_total,
            total_formatted=format_seconds_to_hhmm(user_total)
        ))

    # Build daily totals
    daily_totals = []
    for day in dates_list:
        seconds = daily_totals_seconds.get(day, 0)
        daily_totals.append(TeamTimesheetDayTotal(
            date=day,
            seconds=seconds,
            formatted=format_seconds_to_hhmm(seconds)
        ))

    return TeamTimesheetReport(
        start_date=start_date,
        end_date=end_date,
        dates=dates_list,
        users=timesheet_users,
        daily_totals=daily_totals,
        grand_total_seconds=grand_total,
        grand_total_formatted=format_seconds_to_hhmm(grand_total)
    )


@router.get("/team-timesheet/export/csv")
async def export_team_timesheet_csv(
    start_date: date,
    end_date: date,
    team_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export Team Timesheet report as CSV.
    Returns a downloadable CSV file with user hours per day.
    """
    import csv
    from io import StringIO

    from fastapi.responses import StreamingResponse

    # Get the timesheet data using the same logic
    timesheet = await get_team_timesheet(start_date, end_date, team_id, db, current_user)

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)

    # Header row: Member, Role, Date1, Date2, ..., Total
    header = ["Member", "Role"]
    for d in timesheet.dates:
        header.append(d.strftime("%a %m/%d"))
    header.append("Total")
    writer.writerow(header)

    # Data rows for each user
    for user in timesheet.users:
        row = [user.user_name, user.role]
        for day_entry in user.daily_hours:
            row.append(day_entry.formatted)
        row.append(user.total_formatted)
        writer.writerow(row)

    # Daily totals row
    totals_row = ["Daily Total", ""]
    for day_total in timesheet.daily_totals:
        totals_row.append(day_total.formatted)
    totals_row.append(timesheet.grand_total_formatted)
    writer.writerow(totals_row)

    # Prepare response
    output.seek(0)
    filename = f"team_timesheet_{start_date}_to_{end_date}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/team-timesheet/export/excel")
async def export_team_timesheet_excel(
    start_date: date,
    end_date: date,
    team_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export Team Timesheet report as Excel.
    Returns a downloadable Excel file with formatted user hours per day.
    """
    from io import BytesIO

    from fastapi.responses import StreamingResponse

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available. Install openpyxl.")

    # Get the timesheet data
    timesheet = await get_team_timesheet(start_date, end_date, team_id, db, current_user)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Workbook has no active worksheet"
    ws.title = "Team Timesheet"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    total_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    weekend_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    grand_total_fill = PatternFill(start_color="93C5FD", end_color="93C5FD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(timesheet.dates) + 3)
    title_cell = ws.cell(row=1, column=1, value=f"Team Timesheet: {start_date} to {end_date}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment

    # Header row
    row = 3
    headers = ["Member", "Role"]
    for d in timesheet.dates:
        headers.append(d.strftime("%a\n%m/%d"))
    headers.append("Total")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Data rows
    row = 4
    for user in timesheet.users:
        # Name cell
        name_cell = ws.cell(row=row, column=1, value=user.user_name)
        name_cell.alignment = left_alignment
        name_cell.border = thin_border

        # Role cell
        role_cell = ws.cell(row=row, column=2, value=user.role)
        role_cell.alignment = left_alignment
        role_cell.border = thin_border

        # Daily hours
        col = 3
        for i, day_entry in enumerate(user.daily_hours):
            cell = ws.cell(row=row, column=col, value=day_entry.formatted)
            cell.alignment = center_alignment
            cell.border = thin_border

            # Highlight weekends
            day_date = timesheet.dates[i]
            if day_date.weekday() >= 5:  # Saturday or Sunday
                cell.fill = weekend_fill
            col += 1

        # User total
        total_cell = ws.cell(row=row, column=col, value=user.total_formatted)
        total_cell.alignment = center_alignment
        total_cell.border = thin_border
        total_cell.fill = total_fill
        total_cell.font = Font(bold=True)

        row += 1

    # Daily totals row
    totals_row = row
    ws.cell(row=totals_row, column=1, value="Daily Total").font = Font(bold=True)
    ws.cell(row=totals_row, column=1).border = thin_border
    ws.cell(row=totals_row, column=1).fill = total_fill

    ws.cell(row=totals_row, column=2, value="").border = thin_border
    ws.cell(row=totals_row, column=2).fill = total_fill

    col = 3
    for i, day_total in enumerate(timesheet.daily_totals):
        cell = ws.cell(row=totals_row, column=col, value=day_total.formatted)
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = total_fill
        cell.font = Font(bold=True)
        col += 1

    # Grand total cell
    grand_cell = ws.cell(row=totals_row, column=col, value=timesheet.grand_total_formatted)
    grand_cell.alignment = center_alignment
    grand_cell.border = thin_border
    grand_cell.fill = grand_total_fill
    grand_cell.font = Font(bold=True, size=12)

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    for col_idx in range(3, len(timesheet.dates) + 4):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 26))
        ws.column_dimensions[col_letter].width = 10

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"team_timesheet_{start_date}_to_{end_date}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/team-timesheet/export/pdf")
async def export_team_timesheet_pdf(
    start_date: date,
    end_date: date,
    team_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export Team Timesheet report as PDF.
    Returns a downloadable PDF file with formatted user hours per day.
    """
    from io import BytesIO

    from fastapi.responses import StreamingResponse

    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF export not available. Install reportlab.")

    # Get the timesheet data
    timesheet = await get_team_timesheet(start_date, end_date, team_id, db, current_user)

    # Create PDF in memory
    output = BytesIO()

    # Use landscape orientation for wide tables
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Team Timesheet Report", title_style))

    # Subtitle with date range
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.gray
    )
    elements.append(Paragraph(f"{start_date} to {end_date}", subtitle_style))
    elements.append(Spacer(1, 12))

    # Build table data
    # Header row
    header = ["Member", "Role"]
    for d in timesheet.dates:
        header.append(d.strftime("%a\n%m/%d"))
    header.append("Total")

    table_data = [header]

    # Data rows
    for user in timesheet.users:
        row = [user.user_name, user.role]
        for day_entry in user.daily_hours:
            row.append(day_entry.formatted)
        row.append(user.total_formatted)
        table_data.append(row)

    # Totals row
    totals_row = ["Daily Total", ""]
    for day_total in timesheet.daily_totals:
        totals_row.append(day_total.formatted)
    totals_row.append(timesheet.grand_total_formatted)
    table_data.append(totals_row)

    # Calculate column widths dynamically
    num_cols = len(header)
    page_width = landscape(letter)[0] - 1*inch  # Total width minus margins
    name_col_width = 1.5*inch
    role_col_width = 0.9*inch
    total_col_width = 0.7*inch
    remaining_width = page_width - name_col_width - role_col_width - total_col_width
    date_col_width = remaining_width / (num_cols - 3) if num_cols > 3 else 0.6*inch

    col_widths = [name_col_width, role_col_width]
    col_widths.extend([date_col_width] * (num_cols - 3))
    col_widths.append(total_col_width)

    # Create table
    table = Table(table_data, colWidths=col_widths)

    # Table styling
    style_commands = [
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),

        # Body styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),  # Name and Role left-aligned
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),  # Dates and totals centered
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),

        # Total column styling
        ('BACKGROUND', (-1, 1), (-1, -2), colors.HexColor('#DBEAFE')),
        ('FONTNAME', (-1, 1), (-1, -2), 'Helvetica-Bold'),

        # Totals row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#DBEAFE')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),

        # Grand total cell
        ('BACKGROUND', (-1, -1), (-1, -1), colors.HexColor('#93C5FD')),
        ('FONTNAME', (-1, -1), (-1, -1), 'Helvetica-Bold'),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-2, -2), [colors.white, colors.HexColor('#F9FAFB')]),
    ]

    # Add weekend highlighting (columns where date is Saturday or Sunday)
    for i, d in enumerate(timesheet.dates):
        if d.weekday() >= 5:  # Saturday or Sunday
            col_idx = i + 2  # Offset for Name and Role columns
            style_commands.append(('BACKGROUND', (col_idx, 1), (col_idx, -2), colors.HexColor('#F3F4F6')))

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    # Add summary footer
    elements.append(Spacer(1, 20))
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_LEFT
    )
    elements.append(Paragraph(
        f"<b>Summary:</b> {len(timesheet.users)} team members | "
        f"{len(timesheet.dates)} days | "
        f"Total: {timesheet.grand_total_formatted} hours",
        summary_style
    ))

    # Build PDF
    doc.build(elements)
    output.seek(0)

    filename = f"team_timesheet_{start_date}_to_{end_date}.pdf"

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================
# EMAIL REPORT ENDPOINT
# ============================================

class EmailReportRequest(BaseModel):
    """Schema for email report request"""
    report_type: str  # "time_report", "team_timesheet"
    start_date: date
    end_date: date
    recipients: List[str]  # List of email addresses
    format: str = "pdf"  # "pdf", "excel", "csv"
    custom_message: Optional[str] = None


class EmailReportResponse(BaseModel):
    """Schema for email report response"""
    success: bool
    message: str
    recipients_sent: int
    recipients_failed: int


@router.post("/email", response_model=EmailReportResponse)
async def email_report(
    data: EmailReportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    tz: str = Depends(get_company_timezone),
):
    """
    Send a report via email to specified recipients.

    Supported report types:
    - time_report: Personal or team time report
    - team_timesheet: Team timesheet grid

    Supported formats:
    - pdf: PDF document
    - excel: Excel spreadsheet (.xlsx)
    - csv: CSV file
    """

    from app.services.email_service import email_service

    # Validate admin role for team reports
    if data.report_type == "team_timesheet":
        if current_user.role not in ["admin", "company_admin", "super_admin"]:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only admins can email team timesheets"
            )

    # Validate format
    if data.format not in ["pdf", "excel", "csv"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid format. Must be 'pdf', 'excel', or 'csv'"
        )

    # Validate recipients (basic email format check)
    import re
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    invalid_emails = [e for e in data.recipients if not email_pattern.match(e)]
    if invalid_emails:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid email addresses: {', '.join(invalid_emails)}"
        )

    # Generate report based on type
    try:
        if data.report_type == "time_report":
            attachment_data, attachment_filename, attachment_mimetype = await _generate_time_report(
                db=db,
                user=current_user,
                start_date=data.start_date,
                end_date=data.end_date,
                format=data.format,
                tz=tz,
            )
            report_name = "Time Report"

        elif data.report_type == "team_timesheet":
            attachment_data, attachment_filename, attachment_mimetype = await _generate_team_timesheet_report(
                db=db,
                user=current_user,
                start_date=data.start_date,
                end_date=data.end_date,
                format=data.format,
                tz=tz,
            )
            report_name = "Team Timesheet"

        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unknown report type: {data.report_type}"
            )

    except Exception as e:
        logger.error(f"Failed to generate report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate report: {str(e)}"
        )

    # Format date range for email
    date_range = f"{data.start_date.strftime('%b %d')} - {data.end_date.strftime('%b %d, %Y')}"

    # Send emails
    try:
        results = await email_service.send_report_email(
            to_emails=data.recipients,
            report_name=report_name,
            date_range=date_range,
            attachment_data=attachment_data,
            attachment_filename=attachment_filename,
            attachment_mimetype=attachment_mimetype,
            custom_message=data.custom_message,
            company_id=current_user.company_id,
            db=db
        )

        # Log email results
        for recipient_email, success in results.items():
            if success:
                await log_email_sent(
                    db=db,
                    to_email=recipient_email,
                    subject=f"{report_name} - {date_range}",
                    email_type="report_email",
                    company_id=current_user.company_id,
                    metadata={"report_type": data.report_type, "format": data.format, "date_range": date_range, "user_id": current_user.id}
                )
            else:
                await log_email_failed(
                    db=db,
                    to_email=recipient_email,
                    subject=f"{report_name} - {date_range}",
                    email_type="report_email",
                    error_message="Failed to send report email",
                    company_id=current_user.company_id,
                    metadata={"report_type": data.report_type, "format": data.format, "date_range": date_range, "user_id": current_user.id}
                )

        sent = sum(1 for v in results.values() if v)
        failed = len(results) - sent

        if sent == 0:
            return EmailReportResponse(
                success=False,
                message="Failed to send to all recipients",
                recipients_sent=0,
                recipients_failed=failed
            )
        elif failed > 0:
            return EmailReportResponse(
                success=True,
                message=f"Report sent to {sent} recipient(s), {failed} failed",
                recipients_sent=sent,
                recipients_failed=failed
            )
        else:
            return EmailReportResponse(
                success=True,
                message=f"Report sent successfully to {sent} recipient(s)",
                recipients_sent=sent,
                recipients_failed=0
            )

    except Exception as e:
        logger.error(f"Failed to send report email: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to send email: {str(e)}"
        )


async def _generate_time_report(
    db: AsyncSession,
    user: User,
    start_date: date,
    end_date: date,
    format: str,
    tz: str = "UTC",
) -> tuple:
    """Generate time report and return (data, filename, mimetype).

    ``tz``: IANA timezone for interpreting ``start_date`` / ``end_date`` as
    local civil dates (B7).
    """
    from io import BytesIO

    start_dt, end_dt = range_bounds(start_date, end_date, tz)

    # Get time entries
    query = select(TimeEntry, Project.name.label("project_name"), Task.name.label("task_name")).outerjoin(
        Project, TimeEntry.project_id == Project.id
    ).outerjoin(
        Task, TimeEntry.task_id == Task.id
    ).where(
        TimeEntry.user_id == user.id,
        TimeEntry.start_time >= start_dt,
        TimeEntry.start_time < end_dt,
    ).order_by(TimeEntry.start_time.desc())

    result = await db.execute(query)
    rows = result.all()

    entries = []
    for row in rows:
        entry = row[0]
        duration = 0
        if entry.end_time and entry.start_time:
            duration = int((entry.end_time - entry.start_time).total_seconds())

        entries.append({
            "date": entry.start_time.strftime("%Y-%m-%d"),
            "start_time": entry.start_time.strftime("%H:%M:%S"),
            "end_time": entry.end_time.strftime("%H:%M:%S") if entry.end_time else "Running",
            "duration": format_seconds_to_hhmm(duration),
            "project": row[1] or "No Project",
            "task": row[2] or "No Task",
            "notes": entry.notes or "",
        })

    if format == "csv":
        import csv
        output = BytesIO()
        # CSV needs text wrapper
        import io
        text_output = io.StringIO()
        writer = csv.DictWriter(text_output, fieldnames=["date", "start_time", "end_time", "duration", "project", "task", "notes"])
        writer.writeheader()
        writer.writerows(entries)
        output.write(text_output.getvalue().encode('utf-8'))
        output.seek(0)
        return output.getvalue(), f"time_report_{start_date}_to_{end_date}.csv", "text/csv"

    elif format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        assert ws is not None, "Workbook has no active worksheet"
        ws.title = "Time Report"

        # Headers
        headers = ["Date", "Start Time", "End Time", "Duration", "Project", "Task", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data
        for row_idx, entry in enumerate(entries, 2):
            ws.cell(row=row_idx, column=1, value=entry["date"])
            ws.cell(row=row_idx, column=2, value=entry["start_time"])
            ws.cell(row=row_idx, column=3, value=entry["end_time"])
            ws.cell(row=row_idx, column=4, value=entry["duration"])
            ws.cell(row=row_idx, column=5, value=entry["project"])
            ws.cell(row=row_idx, column=6, value=entry["task"])
            ws.cell(row=row_idx, column=7, value=entry["notes"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue(), f"time_report_{start_date}_to_{end_date}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    else:  # PDF
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph(f"Time Report: {start_date} to {end_date}", styles['Heading1']))
        elements.append(Paragraph(f"User: {user.name}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Table
        table_data = [["Date", "Start", "End", "Duration", "Project", "Task"]]
        for entry in entries:
            table_data.append([
                entry["date"],
                entry["start_time"],
                entry["end_time"],
                entry["duration"],
                entry["project"][:20],
                entry["task"][:20]
            ])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        elements.append(table)

        doc.build(elements)
        output.seek(0)
        return output.getvalue(), f"time_report_{start_date}_to_{end_date}.pdf", "application/pdf"


async def _generate_team_timesheet_report(
    db: AsyncSession,
    user: User,
    start_date: date,
    end_date: date,
    format: str,
    tz: str = "UTC",
) -> tuple:
    """Generate team timesheet report and return (data, filename, mimetype)"""
    from io import BytesIO

    # Get the timesheet data using existing function
    timesheet = await _get_team_timesheet_data(db, user, start_date, end_date, None, tz=tz)

    if format == "csv":
        import csv
        import io

        output = BytesIO()
        text_output = io.StringIO()

        # Build headers
        headers = ["Name", "Role"] + [d.strftime("%m/%d") for d in timesheet.dates] + ["Total"]
        writer = csv.writer(text_output)
        writer.writerow(headers)

        # Write user rows
        for u in timesheet.users:
            row = [u.user_name, u.role]
            for dh in u.daily_hours:
                row.append(dh.formatted if dh.seconds > 0 else "-")
            row.append(u.total_formatted)
            writer.writerow(row)

        # Totals row
        totals_row = ["TOTAL", ""]
        for dt in timesheet.daily_totals:
            totals_row.append(dt.formatted)
        totals_row.append(timesheet.grand_total_formatted)
        writer.writerow(totals_row)

        output.write(text_output.getvalue().encode('utf-8'))
        output.seek(0)
        return output.getvalue(), f"team_timesheet_{start_date}_to_{end_date}.csv", "text/csv"

    elif format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        assert ws is not None, "Workbook has no active worksheet"
        ws.title = "Team Timesheet"

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(timesheet.dates) + 3)
        ws.cell(row=1, column=1, value=f"Team Timesheet: {start_date} to {end_date}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Headers
        headers = ["Name", "Role"] + [d.strftime("%m/%d") for d in timesheet.dates] + ["Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Weekend highlighting
        weekend_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        # Data rows
        for row_idx, u in enumerate(timesheet.users, 4):
            ws.cell(row=row_idx, column=1, value=u.user_name)
            ws.cell(row=row_idx, column=2, value=u.role)

            for col_idx, dh in enumerate(u.daily_hours, 3):
                cell = ws.cell(row=row_idx, column=col_idx, value=dh.formatted if dh.seconds > 0 else "-")
                cell.alignment = Alignment(horizontal="center")
                if dh.date.weekday() >= 5:
                    cell.fill = weekend_fill

            # Total column
            ws.cell(row=row_idx, column=len(timesheet.dates) + 3, value=u.total_formatted)
            ws.cell(row=row_idx, column=len(timesheet.dates) + 3).font = Font(bold=True)

        # Totals row
        totals_row = len(timesheet.users) + 4
        ws.cell(row=totals_row, column=1, value="TOTAL")
        ws.cell(row=totals_row, column=1).font = Font(bold=True)

        for col_idx, dt in enumerate(timesheet.daily_totals, 3):
            cell = ws.cell(row=totals_row, column=col_idx, value=dt.formatted)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws.cell(row=totals_row, column=len(timesheet.dates) + 3, value=timesheet.grand_total_formatted)
        ws.cell(row=totals_row, column=len(timesheet.dates) + 3).font = Font(bold=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue(), f"team_timesheet_{start_date}_to_{end_date}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    else:  # PDF - reuse existing PDF generation logic
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), leftMargin=20, rightMargin=20)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
        elements.append(Paragraph("Team Timesheet", title_style))

        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=12)
        elements.append(Paragraph(f"{start_date.strftime('%B %d')} - {end_date.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Spacer(1, 20))

        # Build table data
        header_row = ["Name", "Role"] + [d.strftime("%m/%d") for d in timesheet.dates] + ["Total"]
        table_data = [header_row]

        for u in timesheet.users:
            row = [u.user_name[:15], u.role[:10]]
            for dh in u.daily_hours:
                row.append(dh.formatted if dh.seconds > 0 else "-")
            row.append(u.total_formatted)
            table_data.append(row)

        # Totals row
        totals_row = ["TOTAL", ""]
        for dt in timesheet.daily_totals:
            totals_row.append(dt.formatted)
        totals_row.append(timesheet.grand_total_formatted)
        table_data.append(totals_row)

        # Calculate column widths
        col_widths = [80, 60] + [35] * len(timesheet.dates) + [45]

        table = Table(table_data, colWidths=col_widths)

        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTNAME', (-1, 0), (-1, -1), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-2, -2), [colors.white, colors.HexColor('#F9FAFB')]),
        ]

        # Weekend highlighting
        for i, d in enumerate(timesheet.dates):
            if d.weekday() >= 5:
                col_idx = i + 2
                style_commands.append(('BACKGROUND', (col_idx, 1), (col_idx, -2), colors.HexColor('#F3F4F6')))

        table.setStyle(TableStyle(style_commands))
        elements.append(table)

        # Summary
        elements.append(Spacer(1, 20))
        summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=10)
        elements.append(Paragraph(
            f"<b>Summary:</b> {len(timesheet.users)} team members | {len(timesheet.dates)} days | Total: {timesheet.grand_total_formatted} hours",
            summary_style
        ))

        doc.build(elements)
        output.seek(0)
        return output.getvalue(), f"team_timesheet_{start_date}_to_{end_date}.pdf", "application/pdf"


async def _get_team_timesheet_data(
    db: AsyncSession,
    user: User,
    start_date: date,
    end_date: date,
    team_id: Optional[int],
    tz: str = "UTC",
) -> TeamTimesheetReport:
    """Helper to get team timesheet data - shared between endpoint and email generator"""
    from app.dependencies import FILTER_NULL_COMPANY, get_company_filter

    now = now_utc()

    # Generate list of all dates in range
    dates_in_range = []
    current = start_date
    while current <= end_date:
        dates_in_range.append(current)
        current += timedelta(days=1)

    # Build user query based on role
    if user.role == "super_admin":
        users_query = select(User).where(User.is_active == True)
    else:
        company_filter = get_company_filter(user)
        if company_filter == FILTER_NULL_COMPANY:
            users_query = select(User).where(User.is_active == True, User.company_id.is_(None))
        else:
            users_query = select(User).where(User.is_active == True, User.company_id == company_filter)

    if team_id:
        users_query = users_query.join(TeamMember, User.id == TeamMember.user_id).where(TeamMember.team_id == team_id)

    users_query = users_query.order_by(User.name)

    result = await db.execute(users_query)
    users = result.scalars().all()

    # Fetch time entries for date range (B7: tenant-local bounds)
    start_datetime, end_datetime = range_bounds(start_date, end_date, tz)

    user_ids = [u.id for u in users]
    if not user_ids:
        return TeamTimesheetReport(
            start_date=start_date,
            end_date=end_date,
            dates=dates_in_range,
            users=[],
            daily_totals=[TeamTimesheetDayTotal(date=d, seconds=0, formatted="0:00") for d in dates_in_range],
            grand_total_seconds=0,
            grand_total_formatted="0:00"
        )

    entries_query = select(TimeEntry).where(
        TimeEntry.user_id.in_(user_ids),
        TimeEntry.start_time < end_datetime,
        ((TimeEntry.end_time >= start_datetime) | (TimeEntry.end_time.is_(None)))
    )

    entries_result = await db.execute(entries_query)
    entries = entries_result.scalars().all()

    # Build user-day matrix
    user_daily_seconds: Dict[int, Dict[date, int]] = defaultdict(lambda: defaultdict(int))

    for entry in entries:
        for d in dates_in_range:
            day_start, day_end = day_bounds(d, tz)
            seconds = calculate_entry_duration_for_period(entry, day_start, day_end, now)
            if seconds > 0:
                user_daily_seconds[entry.user_id][d] += seconds

    # Build response
    timesheet_users = []
    daily_totals_seconds: Dict[date, int] = defaultdict(int)
    grand_total = 0

    for u in users:
        daily_hours = []
        user_total = 0

        for d in dates_in_range:
            seconds = user_daily_seconds[u.id][d]
            daily_hours.append(TeamTimesheetUserEntry(
                date=d,
                seconds=seconds,
                formatted=format_seconds_to_hhmm(seconds) if seconds > 0 else "-"
            ))
            daily_totals_seconds[d] += seconds
            user_total += seconds

        timesheet_users.append(TeamTimesheetUser(
            user_id=u.id,
            user_name=u.name,
            role=u.role,
            daily_hours=daily_hours,
            total_seconds=user_total,
            total_formatted=format_seconds_to_hhmm(user_total)
        ))
        grand_total += user_total

    daily_totals = [
        TeamTimesheetDayTotal(
            date=d,
            seconds=daily_totals_seconds[d],
            formatted=format_seconds_to_hhmm(daily_totals_seconds[d])
        )
        for d in dates_in_range
    ]

    return TeamTimesheetReport(
        start_date=start_date,
        end_date=end_date,
        dates=dates_in_range,
        users=timesheet_users,
        daily_totals=daily_totals,
        grand_total_seconds=grand_total,
        grand_total_formatted=format_seconds_to_hhmm(grand_total)
    )
